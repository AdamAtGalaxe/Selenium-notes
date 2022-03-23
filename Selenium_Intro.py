

import pyautogui
import requests
import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait

def hybrid():
    import openpyxl

    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    path = "/Users/adamroberts/Documents/Selenium/action_sheet.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    nrows = sheet_obj.max_row
    for i in range(2, nrows+1):
        action_type = sheet_obj.cell(row=i, column=2).value
        data = sheet_obj.cell(row=i, column=3).value
        xpath = sheet_obj.cell(row=i, column=4).value

        if action_type == "Open URL":
            try:
                driver.get(data)
                sheet_obj.cell(row=i, column=5).value = "Pass"
            except Exception as e:
                sheet_obj.cell(row=i, column=5).value = "Fail"
                sheet_obj.cell(row=i, column=6).value = e.args[0]
        if action_type == "TEXT":
            try:
                driver.find_element(By.XPATH, xpath).send_keys(data)
                sheet_obj.cell(row=i, column=5).value = "Pass"
            except Exception as e:
                sheet_obj.cell(row=i, column=5).value = "Fail"
                sheet_obj.cell(row=i, column=6).value = e.args[0]
                #sheet_obj.cell(row=i, column=5).font = Font(name="Tahoma", size=12, color="00339966")
                #sheet_obj.cell(row=i, column=5).style = "Hyperlink"

        if action_type == "CLICK" or action_type == "LINK":
            try:
                driver.find_element(By.XPATH, xpath).click()
                sheet_obj.cell(row=i, column=5).value = "Pass"
            except Exception as e:
                sheet_obj.cell(row=i, column=5).value = "Fail"
                sheet_obj.cell(row=i, column=6).value = e.args[0]
        #done on the fly. Not tested
        if action_type == "DROPDOWN":
            try:
                dropdown = Select(xpath)
                dropdown.select_by_visible_text(data)
            except Exception as e:
                sheet_obj.cell(row=i, column=5).value = "Fail"
                sheet_obj.cell(row=i, column=6).value = e.args[0]
            #dropdown.select_by_index(0)
        if action_type == "MouseHover":
            try:
                hover =ActionChains(driver).move_to_element(xpath)
                hover.perform()
            except Exception as e:
                sheet_obj.cell(row=i, column=5).value = "Fail"
                sheet_obj.cell(row=i, column=6).value = e.args[0]

    wb_obj.save("/Users/adamroberts/Documents/Selenium/action_sheet.xlsx")
    driver.quit()
"""
driver.execute_script("arguments[0].scrollIntoView();", element)

time.sleep(15)
w=WebDriverWait(driver, 15)
w.until(expected_condition)

alb.web_click(fb.byLocator(pobj.btn_promptbutton(), "Prompt Button"))

driver.find_element(By.XPATH, "//button{@id='promptButton']").click()

try:
    w = WebDriverWait(driver, 15)
    w.until(expected_condition.alert_is_present(fb.byLocator(pobj.btn_promptbutton())))
alb.manage_alert()
time.sleep(5)


def byLocatior(element_type, elem_str):
    if element_type == "XPATH":
        element = driver.find_element(By.XPATH, elem_str)
    if element_type == "Name":
        element = driver.find_element(By.NAME, elem_str)
    return element

element = byLocator(driver.find_element(By.XPATH, elem_str)

"""
def frames():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://demoqa.com/frames"
    driver.get(url)
    #driver.switch_to.
    driver.switch_to.frame("frame1")
    extracted_text = driver.find_element(By.XPATH, "//h1[contains(text(), 'This is a')][1]").text
    print(extracted_text)

def nestedFrames():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://demoqa.com/nestedframes"
    driver.get(url)
    driver.switch_to.frame("frame1")
    element = driver.find_element(By.XPATH,"//iframe[@srcdoc = '<p>Child Iframe</p>']")
    driver.switch_to.frame(element)
    extracted_text = driver.find_element(By.XPATH, "//p[text() = 'Child Iframe']").text
    print(extracted_text)


def pyautoguiTAB():  #GET HEEEEEEEEEEEEELLLLLLLLLLLLLLLP!!!!!
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://demoqa.com/automation-practice-form"
    driver.get(url)
    driver.find_element(By.XPATH, "//input[@id='firstName']").send_keys("Adam")
    time.sleep(1)
    pyautogui.press('TAB')
    time.sleep(2)
    pyautogui.press('TAB')
    time.sleep(2)
    pyautogui.press('TAB')

def screenshot():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://flipkart.com"
    driver.get(url)
    driver.save_screenshot("/Users/adam/Documents/rando/my_screenshot.png") #"C:\windowsDirectory\my_screenshot.png"

def actionChains():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://flipkart.com"
    driver.get(url)
    time.sleep(10)
    actions = ActionChains(driver)
    element = driver.find_element(By.XPATH, "//img[@alt='Electronics']")
    actions.move_to_element(element)
    """
    actions.click(element)
    actions.double_click(element)
    etc
    """
    actions.perform()
    time.sleep(10)

def isDisplayed():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://flipkart.com"
    driver.get(url)
    time.sleep(2)
    if driver.find_element(By.XPATH, "//img[@alt='Flipkart']").is_displayed():
        print("FlipKart is opened")
    else:
        print("nope")

def tryCatch():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    try:
        url = "https://flipkart.com"
        driver.get(url)
        time.sleep(2)
        if driver.find_element(By.XPATH, "//img[@alt='Flipkart2']").is_displayed():
            print("FlipKart is opened")
        else:
            print("nope")
    except Exception as e:
        print("There is an error")\
        #print(str(e))

def excel():
    import xlrd

    try:
        wb = xlrd.open_workbook("/Users/adam/Documents/rando/my_data.xls")
        sheet = wb.sheet_by_index(0)
        cnt_row = sheet.nrows
        cnt_col = sheet.ncols
        print("Excel")

        for i in range(1, cnt_row):
            name = sheet.cell_value(i, 0)
            email = sheet.cell_value(i, 1)
            password = sheet.cell_value(i, 2)
            print("name: {}, email: {}, password: {}".format(name, email, password))
    except Exception as e:
        print("error in reading excel")
        print(str(e))

def wait():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://demoqa.com/alerts"
    driver.get(url)
    driver.maximize_window()

    # implicit wait/between each call/fix race conditions
    driver.implicitly_wait(5)
    element = driver.find_element(By.XPATH, "//*[@id='promtButton']")
    driver.execute_script("arguments[0].scrollIntoView();", element)

    # hard sleep
    time.sleep(5)

    # explicit wait/smart wait
    w = WebDriverWait(driver, 10)
    w.until(expected_conditions.element_to_be_clickable((By.XPATH, "//*[@id='promtButton']")))
    driver.find_element(By.XPATH, "//*[@id='promtButton']").click()
    driver.switch_to.alert.send_keys('test')

def alerts():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://demoqa.com/alerts"
    driver.get(url)
    driver.find_element(By.XPATH,"//*[@id='alertButton']").click()
    time.sleep(5)
    alert = driver.switch_to.alert
    alert.accept()
    alert.dismiss()
    time.sleep(5)

def promptBox():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://demoqa.com/alerts"
    driver.get(url)
    element = driver.find_element(By.XPATH, "//*[@id='promtButton']")
    driver.execute_script("arguments[0].scrollIntoView();", element)
    element.click()
    time.sleep(1)
    alert2 = driver.switch_to.alert

    alert2.send_keys("test")
    #alert2.send_keys(keys.RETURN)
    #alert2.accept()
    #alert.send_keys('test')
    #alert.c # alert.dismiss()
    time.sleep(5)

def tabSwitch():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://www.facebook.com"
    driver.get(url)

    for i in range(1, 6):
        driver.switch_to.new_window()

    handles = driver.window_handles
    len_handles = len(handles)

    for i in range(1,len_handles):
        driver.switch_to.window(handles[i])
        time.sleep(2)

def statusCode():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    #This is problematic..
    url = "https://www.facebook.com"
    r = requests.get(url)
    print(r.status_code)


def noRightClick():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://www.facebook.com"
    driver.get(url)
    page_source = driver.page_source
    print(page_source)

    if "fldLoginUserID" in page_source:
        print("got element")
    else:
        print("Nope")



def getURL():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://www.facebook.com"
    driver.get(url)
    driver.find_element(By.XPATH, "//*[@id='email']").send_keys("adamrobertsburner@gmail.com")
    driver.find_element(By.XPATH, "//*[@id='pass']").send_keys("Burner4u!")
    driver.find_element(By.XPATH, "//*[@name='login']").click()
    print(driver.current_url)


def pageTitle():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://www.facebook.com"
    driver.get(url)
    driver.find_element(By.XPATH, "//*[@id='email']").send_keys("adamrobertsburner@gmail.com")
    driver.find_element(By.XPATH, "//*[@id='pass']").send_keys("Burner4u!")
    driver.find_element(By.XPATH, "//*[@name='login']").click()
    page_title = driver.title
    print(page_title)
    if "Facebook" in page_title:
        print("You are on facebook")
    else:
        print("NOPE: " + driver.title)

#busted/Will fix
def findElements():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    url = "https://www.facebook.com"
    driver.get(url)
    mylist = driver.find_elements(By.XPATH, "//input")
    print(len(mylist))

    for i in mylist:
        print(i)
        # driver.find_elements_by_xpath("(//input)" + i +"").send_keys("test")
        # driver.find_element(By.XPATH, "(//input)" + i +"").send_keys("test")

def basicNavigation():
    s = Service("/Users/adamroberts/Documents/Selenium/chromedriver")
    driver = webdriver.Chrome(service=s)
    driver.get("https://gmail.com")
    driver.maximize_window()
    driver.get("https://www.facebook.com")
    driver.find_element(By.XPATH, "//*[@id='email']").send_keys("adamrobertsburner@gmail.com")
    driver.find_element(By.XPATH, "//*[@id='pass']").send_keys("Burner4u!")
    driver.find_element(By.XPATH, "//*[@name='login']").click()
    driver.back()
    driver.forward()
    driver.quit()
